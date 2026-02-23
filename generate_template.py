#!/usr/bin/env python3
"""Generate the trial balance import template for StatementHub."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Sheet 1: Trial Balance ─────────────────────────────────────────
ws = wb.active
ws.title = "Trial Balance"

# Styles
header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
sample_font = Font(name="Calibri", size=10, color="999999", italic=True)
thin_border = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
number_format = '#,##0.00'

# Headers — Opening Balance removed
headers = ["Account Code", "Account Name", "Debit", "Credit"]
col_widths = [18, 45, 20, 20]

for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Sample data row (row 2) — italic grey to indicate "replace me"
sample_data = ["1000", "Cash at Bank - CBA", 125000.00, 98500.00]
for col_idx, value in enumerate(sample_data, start=1):
    cell = ws.cell(row=2, column=col_idx, value=value)
    cell.font = sample_font
    cell.border = thin_border
    if col_idx >= 3:
        cell.number_format = number_format
        cell.alignment = Alignment(horizontal="right", vertical="center")
    else:
        cell.alignment = Alignment(horizontal="left", vertical="center")

# Bordered empty rows from 3 to 1001 (1,000 data rows total including sample)
for row_idx in range(3, 1002):
    for col_idx in range(1, 5):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        if col_idx >= 3:
            cell.number_format = number_format
            cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

# Freeze header row
ws.freeze_panes = "A2"

# Print settings
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

# ── Sheet 2: Instructions ─────────────────────────────────────────
wi = wb.create_sheet(title="Instructions")

title_font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
heading_font = Font(name="Calibri", bold=True, size=12, color="1F4E79")
body_font = Font(name="Calibri", size=11)
bold_font = Font(name="Calibri", bold=True, size=11)
note_font = Font(name="Calibri", size=11, color="C0392B")

wi.column_dimensions["A"].width = 4
wi.column_dimensions["B"].width = 25
wi.column_dimensions["C"].width = 80

row = 1
wi.merge_cells("A1:C1")
cell = wi.cell(row=row, column=1, value="StatementHub — Trial Balance Import Template")
cell.font = title_font
row += 2

# Section 1: Overview
cell = wi.cell(row=row, column=1, value="Overview")
cell.font = heading_font
row += 1
cell = wi.cell(row=row, column=1,
    value="Use the 'Trial Balance' sheet to enter your trial balance data. "
          "Enter the total debit and credit movements for each account during the financial year. "
          "The first row contains headers — do not modify or delete them. "
          "Replace the grey sample row with your actual data.")
cell.font = body_font
cell.alignment = Alignment(wrap_text=True)
wi.merge_cells(f"A{row}:C{row}")
row += 2

# Section 2: Column Descriptions
cell = wi.cell(row=row, column=1, value="Column Descriptions")
cell.font = heading_font
row += 1

columns_info = [
    ("Account Code", "Your chart of accounts code (e.g., 1000, 2100, 4000). "
                     "Text or numeric. Must be unique per row."),
    ("Account Name", "Descriptive name for the account (e.g., 'Cash at Bank - CBA', "
                     "'Trade Debtors'). Free text."),
    ("Debit", "Total debit movements for the period. Enter as a positive number. "
              "Leave blank or enter 0 if there are no debit movements."),
    ("Credit", "Total credit movements for the period. Enter as a positive number. "
               "Leave blank or enter 0 if there are no credit movements."),
]

for col_name, description in columns_info:
    cell = wi.cell(row=row, column=2, value=col_name)
    cell.font = bold_font
    cell = wi.cell(row=row, column=3, value=description)
    cell.font = body_font
    cell.alignment = Alignment(wrap_text=True)
    row += 1

row += 1

# Section 3: Important Notes
cell = wi.cell(row=row, column=1, value="Important Notes")
cell.font = heading_font
row += 1

notes = [
    "The file must be saved as .xlsx format (Excel 2007+).",
    "Row 1 must contain the headers exactly as shown. Do not rename or reorder columns.",
    "Delete the grey sample row before importing — it is provided as a format guide only.",
    "Enter only the movements (Debit and Credit) for the financial year — not opening or closing balances.",
    "Both P&L accounts (income, expenses) and Balance Sheet accounts (assets, liabilities, equity) should be included.",
    "Debits and credits should balance (total debits = total credits) for a valid trial balance.",
    "Importing will replace all existing non-adjustment trial balance lines for the financial year.",
    "Account codes will be automatically mapped to financial statement line items where possible.",
    "Unmapped accounts can be manually mapped after import via the Trial Balance screen.",
    "Up to 1,000 accounts can be imported per file. Maximum file size: 20 MB.",
]

for i, note in enumerate(notes, start=1):
    cell = wi.cell(row=row, column=2, value=f"{i}.")
    cell.font = bold_font
    cell = wi.cell(row=row, column=3, value=note)
    cell.font = body_font
    cell.alignment = Alignment(wrap_text=True)
    row += 1

row += 1
wi.merge_cells(f"A{row}:C{row}")
cell = wi.cell(row=row, column=1,
    value="If you encounter errors during import, check that all numeric columns "
          "contain valid numbers and that no rows have missing Account Codes.")
cell.font = note_font
cell.alignment = Alignment(wrap_text=True)

# Save
output_path = "static/trial_balance_template.xlsx"
wb.save(output_path)
print(f"Template saved to {output_path}")
print(f"Columns: {headers}")
print(f"Data rows: 1000 bordered rows (rows 2-1001)")
