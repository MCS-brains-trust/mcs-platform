"""
Generate the bank statement upload template for StatementHub.
Creates an Excel file with:
- An Instructions sheet
- A Bank Statement sheet with metadata rows and column headers
- Column A (Date) from row 7 onwards formatted as DD/MM/YY so pasted
  dates display uniformly regardless of source format
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──────────────────────────────────────────────────────────────
header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
meta_label_font = Font(name='Calibri', bold=True, size=11)
meta_value_font = Font(name='Calibri', size=11)
meta_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
instruction_font = Font(name='Calibri', size=11)
title_font = Font(name='Calibri', bold=True, size=14, color='1B5E20')
subtitle_font = Font(name='Calibri', bold=True, size=12, color='37474F')
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

# ── Instructions Sheet ──────────────────────────────────────────────────
ws_instr = wb.active
ws_instr.title = 'Instructions'
ws_instr.sheet_properties.tabColor = '1B5E20'

instructions = [
    ('StatementHub — Bank Statement Upload Template', title_font, None),
    ('', None, None),
    ('How to use this template', subtitle_font, None),
    ('', None, None),
    ('1. Go to the "Bank Statement" tab in this workbook.', instruction_font, None),
    ('2. Fill in the metadata section (rows 1-4) with your bank account details:', instruction_font, None),
    ('   - Opening Balance: The balance at the start of the period', instruction_font, None),
    ('   - Closing Balance: The balance at the end of the period', instruction_font, None),
    ('   - BSB: Your bank\'s BSB number (e.g. 063-123)', instruction_font, None),
    ('   - Account Number: Your bank account number', instruction_font, None),
    ('', None, None),
    ('3. Copy and paste your bank transactions starting from row 7.', instruction_font, None),
    ('   Dates will automatically display as DD/MM/YY.', instruction_font, None),
    ('', None, None),
    ('   Column A: Date', instruction_font, None),
    ('   Column B: Description', instruction_font, None),
    ('   Column C: Amount (negative = withdrawal/debit, positive = deposit/credit)', instruction_font, None),
    ('   Column D: Balance (optional)', instruction_font, None),
    ('', None, None),
    ('Important Notes', subtitle_font, None),
    ('', None, None),
    ('• Withdrawals (debits) should be negative values (e.g. -150.00).', instruction_font, None),
    ('• Deposits (credits) should be positive values (e.g. 500.00).', instruction_font, None),
    ('• The Opening Balance must match the trial balance for this bank account.', instruction_font, None),
    ('• Dates must fall within the financial year period selected during upload.', instruction_font, None),
    ('• Do not modify the column headers in row 6.', instruction_font, None),
    ('• Do not add extra columns or merge cells.', instruction_font, None),
]

for i, (text, font, fill) in enumerate(instructions, 1):
    cell = ws_instr.cell(row=i, column=1, value=text)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill

ws_instr.column_dimensions['A'].width = 80

# ── Bank Statement Sheet ────────────────────────────────────────────────
ws = wb.create_sheet('Bank Statement')
ws.sheet_properties.tabColor = '2E7D32'

# Metadata section (rows 1-4)
meta_labels = [
    ('Opening Balance', '0.00'),
    ('Closing Balance', '0.00'),
    ('BSB', ''),
    ('Account Number', ''),
]

for i, (label, default) in enumerate(meta_labels, 1):
    label_cell = ws.cell(row=i, column=1, value=label)
    label_cell.font = meta_label_font
    label_cell.fill = meta_fill
    label_cell.border = thin_border
    label_cell.alignment = Alignment(horizontal='right')

    value_cell = ws.cell(row=i, column=2, value=default)
    value_cell.font = meta_value_font
    value_cell.fill = meta_fill
    value_cell.border = thin_border
    if i <= 2:  # Opening/Closing balance — number format
        value_cell.number_format = '#,##0.00'

# Row 5: blank separator
ws.cell(row=5, column=1, value='')

# Row 6: Column headers
headers = ['Date', 'Description', 'Amount', 'Balance']
col_widths = [14, 50, 18, 15]

for j, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=6, column=j, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')
    ws.column_dimensions[get_column_letter(j)].width = width

# ── Pre-format data rows (7 to 5006) ──────────────────────────────────
# Column A: DD/MM/YY date format — when a date value (or Excel serial
# number) is pasted, Excel will display it as DD/MM/YY.
# Columns C-D: Number format for currency amounts.
for i in range(7, 5007):
    # Date column — DD/MM/YY display format
    date_cell = ws.cell(row=i, column=1)
    date_cell.number_format = 'DD/MM/YY'

    # Amount, Balance — number format
    for j in range(3, 5):
        num_cell = ws.cell(row=i, column=j)
        num_cell.number_format = '#,##0.00'

# Freeze panes at row 7 (below headers)
ws.freeze_panes = 'A7'

# Set the Bank Statement sheet as active
wb.active = wb.sheetnames.index('Bank Statement')

# Save
output_path = 'static/bank_statement_template.xlsx'
wb.save(output_path)
print(f'Template saved to {output_path}')
