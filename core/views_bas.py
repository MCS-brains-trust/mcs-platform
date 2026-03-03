"""
BAS Period-Aware Views.

Replaces the monolithic gst_activity_statement view with a period-aware
architecture. The main entry point is `bas_dashboard` which renders the
full BAS page with period selector, status strip, and period-filtered
GST calculations.

Additional endpoints handle lodgement, unlodgement, and document downloads.
"""
import io
from datetime import date
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font, numbers

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from .models import (
    BASPeriod, Entity, FinancialYear, AuditLog,
    ChartOfAccount, EntityChartOfAccount,
)
from .bas_utils import (
    ensure_bas_periods,
    get_bank_coverage,
    calculate_gst_for_period,
    compute_period_status,
)
from config.authorization import get_financial_year_for_user


def _log_action(request, action, description, obj=None):
    """Create an audit log entry."""
    AuditLog.objects.create(
        user=request.user,
        action=action,
        description=description,
        affected_object_type=type(obj).__name__ if obj else "",
        affected_object_id=str(obj.pk) if obj else "",
        ip_address=request.META.get("REMOTE_ADDR"),
    )


# ── Main BAS Dashboard ──────────────────────────────────────────────────────

@login_required
def bas_dashboard(request, pk):
    """
    Period-aware GST/BAS dashboard.

    Renders the full BAS page with:
      - Period selector bar (Zone B)
      - Period status strip (Zone C)
      - Summary cards (Zone D) — filtered to selected period
      - Detail workspace (Zone E) — filtered to selected period
    """
    fy = get_financial_year_for_user(request, pk)
    fy = get_object_or_404(
        FinancialYear.objects.select_related("entity", "entity__client"),
        pk=pk,
    )
    entity = fy.entity

    # Determine period type from entity's BAS frequency
    period_type = getattr(entity, "bas_frequency", "quarterly") or "quarterly"

    # Ensure all BASPeriod records exist
    periods = ensure_bas_periods(fy, period_type)

    # Build period data with dynamic status and coverage
    period_data = []
    for bp in periods:
        dynamic_status = compute_period_status(fy, bp.period_start, bp.period_end, bp)
        # Update status in DB if it changed (but don't overwrite 'lodged')
        if bp.status != "lodged" and bp.status != dynamic_status:
            bp.status = dynamic_status
            bp.save(update_fields=["status"])

        coverage = get_bank_coverage(fy, bp.period_start, bp.period_end)
        period_data.append({
            "period": bp,
            "status": bp.status,
            "coverage": coverage,
        })

    # Determine which period to show (from query param or auto-select)
    selected_period_num = request.GET.get("period")
    selected_view = request.GET.get("view", "")  # "full_year" or ""

    if selected_view == "full_year":
        # Full Year view
        selected_period = None
        period_start = None
        period_end = None
    elif selected_period_num:
        try:
            num = int(selected_period_num)
            selected_period = next(
                (pd for pd in period_data if pd["period"].period_number == num), None
            )
        except (ValueError, StopIteration):
            selected_period = None

        if selected_period:
            period_start = selected_period["period"].period_start
            period_end = selected_period["period"].period_end
        else:
            selected_period = None
            period_start = None
            period_end = None
    else:
        # Auto-select: earliest 'ready', then 'partial', then full year if all lodged, then first period
        selected_period = None
        for pd in period_data:
            if pd["status"] == "ready":
                selected_period = pd
                break
        if not selected_period:
            for pd in period_data:
                if pd["status"] == "partial":
                    selected_period = pd
                    break
        if not selected_period:
            all_lodged = all(pd["status"] == "lodged" for pd in period_data)
            if all_lodged and period_data:
                # Full Year view
                period_start = None
                period_end = None
            elif period_data:
                selected_period = period_data[0]

        if selected_period:
            period_start = selected_period["period"].period_start
            period_end = selected_period["period"].period_end
        elif selected_view != "full_year":
            # Default to full year if nothing selected
            period_start = None
            period_end = None
            selected_view = "full_year"

    # Calculate GST figures for the selected period (or full year)
    gst_result = calculate_gst_for_period(fy, period_start, period_end)

    # Build period label for display
    if selected_period:
        period_label = selected_period["period"].label
        period_date_range = (
            f"{selected_period['period'].period_start.strftime('%d %b %Y')} — "
            f"{selected_period['period'].period_end.strftime('%d %b %Y')}"
        )
    else:
        period_label = "Full Year"
        period_date_range = (
            f"{fy.start_date.strftime('%d %b %Y')} — "
            f"{fy.end_date.strftime('%d %b %Y')}"
        )

    # Status colours for template
    status_colours = {
        "lodged": "#2E8B57",
        "ready": "#2E75B6",
        "partial": "#E8A317",
        "empty": "#999999",
    }

    # ── Build grouped transaction detail for Sales/Purchases tabs ──
    sales_txns = gst_result.get("sales_transactions", [])
    purchase_txns = gst_result.get("purchase_transactions", [])

    def _group_transactions(txns):
        """Group transactions by GST status (GST vs GST-Free/BAS Excluded)."""
        gst_txns = [t for t in txns if t.get("has_gst")]
        non_gst_txns = [t for t in txns if not t.get("has_gst")]

        gst_total_taxable = sum(t["taxable_amount"] for t in gst_txns)
        gst_total_gst = sum(t["gst_amount"] for t in gst_txns)
        gst_total_gross = sum(t["gross_amount"] for t in gst_txns)

        non_gst_total = sum(t["gross_amount"] for t in non_gst_txns)

        return {
            "gst_txns": gst_txns,
            "non_gst_txns": non_gst_txns,
            "gst_total_taxable": gst_total_taxable,
            "gst_total_gst": gst_total_gst,
            "gst_total_gross": gst_total_gross,
            "non_gst_total": non_gst_total,
            "grand_total": gst_total_gross + non_gst_total,
        }

    sales_detail = _group_transactions(sales_txns)
    purchase_detail = _group_transactions(purchase_txns)

    context = {
        "fy": fy,
        "entity": entity,
        "period_type": period_type,
        "periods": period_data,
        "selected_period": selected_period,
        "selected_view": selected_view,
        "period_label": period_label,
        "period_date_range": period_date_range,
        "bas_data": gst_result["bas_data"],
        "sales_lines": gst_result["sales_lines"],
        "purchase_lines": gst_result["purchase_lines"],
        "capital_lines": gst_result["capital_lines"],
        "excluded_lines": gst_result["excluded_lines"],
        "is_gst_registered": entity.is_gst_registered,
        "status_colours": status_colours,
        "sales_detail": sales_detail,
        "purchase_detail": purchase_detail,
    }
    return render(request, "core/gst_activity_statement.html", context)


# ── Lodgement Endpoints ─────────────────────────────────────────────────────

@login_required
@require_POST
def bas_lodge_period(request, pk, period_number):
    """
    Mark a BAS period as lodged. Captures a snapshot of the GST figures
    at the time of lodgement for audit purposes.
    """
    fy = get_financial_year_for_user(request, pk)
    entity = fy.entity
    period_type = getattr(entity, "bas_frequency", "quarterly") or "quarterly"

    if not request.user.can_do_accounting:
        messages.error(request, "You do not have permission to lodge BAS periods.")
        return redirect("core:gst_activity_statement", pk=pk)

    periods = ensure_bas_periods(fy, period_type)
    bp = get_object_or_404(BASPeriod, financial_year=fy, period_type=period_type, period_number=period_number)

    if bp.status == "lodged":
        messages.warning(request, f"{bp.label} is already lodged.")
        return redirect("core:gst_activity_statement", pk=pk)

    # Check coverage
    coverage = get_bank_coverage(fy, bp.period_start, bp.period_end)
    override_reason = request.POST.get("override_reason", "").strip()

    if coverage["status"] != "complete" and not override_reason:
        messages.error(
            request,
            f"Cannot lodge {bp.label}: incomplete bank coverage. "
            f"Missing months: {', '.join(coverage['missing'])}. "
            f"Provide an override reason to proceed."
        )
        return redirect("core:gst_activity_statement", pk=pk)

    # Calculate GST snapshot
    gst_result = calculate_gst_for_period(fy, bp.period_start, bp.period_end)
    bas = gst_result["bas_data"]

    # Update the period
    bp.status = "lodged"
    bp.lodged_by = request.user
    bp.lodged_at = timezone.now()
    bp.snapshot_1a = bas["1A"]
    bp.snapshot_1b = bas["1B"]
    bp.snapshot_net = bas["gst_payable"]
    if override_reason:
        bp.override_reason = override_reason
    bp.save()

    _log_action(
        request, "lodge",
        f"Lodged BAS {bp.label} for {entity.entity_name} ({fy.year_label}). "
        f"1A={bas['1A']}, 1B={bas['1B']}, Net={bas['gst_payable']}",
        bp,
    )

    messages.success(request, f"{bp.label} marked as lodged.")
    return redirect(f"{request.META.get('HTTP_REFERER', '')}") or redirect(
        "core:gst_activity_statement", pk=pk
    )


@login_required
@require_POST
def bas_unlodge_period(request, pk, period_number):
    """
    Unlodge a BAS period. Restricted to reviewers and admins.
    Preserves the original lodgement data for audit trail.
    """
    fy = get_financial_year_for_user(request, pk)
    entity = fy.entity
    period_type = getattr(entity, "bas_frequency", "quarterly") or "quarterly"

    if not request.user.is_senior:
        messages.error(request, "Only senior accountants and administrators can unlodge BAS periods.")
        return redirect("core:gst_activity_statement", pk=pk)

    bp = get_object_or_404(BASPeriod, financial_year=fy, period_type=period_type, period_number=period_number)

    if bp.status != "lodged":
        messages.warning(request, f"{bp.label} is not currently lodged.")
        return redirect("core:gst_activity_statement", pk=pk)

    reason = request.POST.get("unlodge_reason", "").strip()
    if not reason:
        messages.error(request, "A reason is required to unlodge a BAS period.")
        return redirect("core:gst_activity_statement", pk=pk)

    # Unlodge — the snapshot fields are preserved for audit trail
    bp.status = "ready"  # Will be recalculated dynamically
    bp.unlodged_by = request.user
    bp.unlodged_at = timezone.now()
    bp.save()

    _log_action(
        request, "unlodge",
        f"Unlodged BAS {bp.label} for {entity.entity_name} ({fy.year_label}). "
        f"Reason: {reason}. Original snapshot: 1A={bp.snapshot_1a}, 1B={bp.snapshot_1b}, Net={bp.snapshot_net}",
        bp,
    )

    messages.success(request, f"{bp.label} has been unlodged. Reason: {reason}")
    return redirect("core:gst_activity_statement", pk=pk)


# ── Coverage API ─────────────────────────────────────────────────────────────

@login_required
def bas_coverage_check(request, pk, period_number):
    """
    Return bank coverage details for a specific BAS period.
    Used by the frontend to show coverage indicators.
    """
    fy = get_financial_year_for_user(request, pk)
    entity = fy.entity
    period_type = getattr(entity, "bas_frequency", "quarterly") or "quarterly"

    bp = get_object_or_404(BASPeriod, financial_year=fy, period_type=period_type, period_number=period_number)
    coverage = get_bank_coverage(fy, bp.period_start, bp.period_end)

    return JsonResponse({
        "period": bp.period_number,
        "label": bp.label,
        "status": coverage["status"],
        "months": coverage["months"],
        "missing": coverage["missing"],
    })


# ── Document Download (Period-Scoped) ────────────────────────────────────────

@login_required
def bas_download(request, pk):
    """
    Download GST Activity Statement as Excel or PDF.
    Supports period-scoped downloads and "All Periods" multi-sheet export.
    """
    fmt = request.GET.get("format", "excel").lower()
    period_num = request.GET.get("period")
    view_type = request.GET.get("view", "")

    fy = get_financial_year_for_user(request, pk)
    fy = get_object_or_404(
        FinancialYear.objects.select_related("entity", "entity__client"),
        pk=pk,
    )
    entity = fy.entity
    period_type = getattr(entity, "bas_frequency", "quarterly") or "quarterly"

    if view_type == "all_periods":
        return _download_all_periods(fy, entity, period_type, fmt)

    # Determine period scope
    if period_num and view_type != "full_year":
        try:
            num = int(period_num)
            bp = BASPeriod.objects.get(
                financial_year=fy, period_type=period_type, period_number=num
            )
            period_start = bp.period_start
            period_end = bp.period_end
            period_label = bp.label
        except (ValueError, BASPeriod.DoesNotExist):
            period_start = None
            period_end = None
            period_label = "Full Year"
    else:
        period_start = None
        period_end = None
        period_label = "Full Year"

    gst_result = calculate_gst_for_period(fy, period_start, period_end)
    bas = gst_result["bas_data"]

    # Build detail rows
    detail_rows = []
    for lines in [gst_result["sales_lines"], gst_result["purchase_lines"], gst_result["capital_lines"]]:
        for row in lines:
            detail_rows.append({
                "code": row["code"],
                "name": row["name"],
                "tax_code": row.get("tax_code", "N-T"),
                "amount": float(row["amount"]),
                "bas_label": row.get("bas_label", ""),
            })

    if fmt == "pdf":
        return _generate_pdf(fy, entity, period_label, bas, detail_rows, period_start, period_end)

    return _generate_excel(fy, entity, period_label, bas, detail_rows, period_start, period_end)


def _generate_excel(fy, entity, period_label, bas, detail_rows, period_start=None, period_end=None):
    """Generate a single-period Excel workbook."""
    wb = openpyxl.Workbook()
    _write_bas_sheet(wb.active, fy, entity, period_label, bas, detail_rows, period_start, period_end)

    # Detail sheet
    ws2 = wb.create_sheet("Detail Breakdown")
    _write_detail_sheet(ws2, detail_rows)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_name = entity.entity_name.replace(" ", "_")
    safe_period = period_label.replace(" ", "_").replace("(", "").replace(")", "")
    if period_start and period_end:
        filename = f"BAS_{safe_name}_{safe_period}_{period_start.strftime('%Y%m%d')}_{period_end.strftime('%Y%m%d')}.xlsx"
    else:
        filename = f"BAS_{safe_name}_FullYear_{fy.start_date.strftime('%Y%m%d')}_{fy.end_date.strftime('%Y%m%d')}.xlsx"

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _generate_pdf(fy, entity, period_label, bas, detail_rows, period_start=None, period_end=None):
    """Generate a single-period PDF."""
    import weasyprint
    from django.utils.html import escape as html_escape

    abn = html_escape(entity.abn or 'N/A')
    if period_start and period_end:
        period_str = f"{period_label}: {period_start.strftime('%d/%m/%Y')} to {period_end.strftime('%d/%m/%Y')}"
    else:
        period_str = f"Full Year: {fy.start_date.strftime('%d/%m/%Y')} to {fy.end_date.strftime('%d/%m/%Y')}"

    net_gst = bas["gst_payable"]
    net_label = "Net GST Payable to ATO" if net_gst > 0 else "Net GST Refundable from ATO"

    def fmt(v):
        return f"${v:,.0f}"

    # Build detail rows HTML
    detail_html = ""
    for row in detail_rows:
        detail_html += f"""<tr>
            <td>{html_escape(str(row['code']))}</td><td>{html_escape(str(row['name']))}</td>
            <td>{html_escape(str(row['tax_code']))}</td><td class="r">${row['amount']:,.0f}</td>
            <td>{html_escape(str(row['bas_label']))}</td>
        </tr>"""

    # Lodgement status footer
    bp = None
    if period_start:
        period_type = getattr(entity, "bas_frequency", "quarterly") or "quarterly"
        bp = BASPeriod.objects.filter(
            financial_year=fy, period_type=period_type,
            period_start=period_start, period_end=period_end,
        ).first()

    lodgement_footer = ""
    if bp and bp.status == "lodged":
        lodgement_footer = f"""
        <div style="margin-top: 10mm; padding: 3mm; background: #e8f5e9; border: 1px solid #2E8B57; font-size: 9pt;">
            <strong>LODGED</strong> by {html_escape(str(bp.lodged_by))} on {bp.lodged_at.strftime('%d/%m/%Y %H:%M') if bp.lodged_at else 'N/A'}
            | Snapshot: 1A={fmt(bp.snapshot_1a or 0)}, 1B={fmt(bp.snapshot_1b or 0)}, Net={fmt(bp.snapshot_net or 0)}
        </div>"""

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
    @page {{ size: A4; margin: 15mm; }}
    body {{ font-family: 'Times New Roman', serif; font-size: 10pt; color: #333; }}
    h1 {{ font-size: 14pt; text-align: center; margin-bottom: 2mm; }}
    h2 {{ font-size: 11pt; margin-top: 6mm; margin-bottom: 3mm; border-bottom: 1px solid #333; padding-bottom: 2mm; }}
    .sub {{ text-align: center; font-size: 10pt; color: #666; margin-bottom: 4mm; }}
    table {{ width: 100%; border-collapse: collapse; margin-bottom: 4mm; }}
    th, td {{ padding: 3px 6px; font-size: 9pt; border-bottom: 1px solid #ddd; }}
    th {{ background: #f5f5f5; text-align: left; font-weight: bold; }}
    .r {{ text-align: right; }}
    .bold {{ font-weight: bold; }}
    .highlight {{ background: #e8f4fd; }}
    .summary {{ background: #f8f9fa; }}
    .total-row {{ border-top: 2px solid #333; font-weight: bold; }}
</style></head><body>

<h1>GST Activity Statement</h1>
<p class="sub">{html_escape(entity.entity_name)} &mdash; ABN: {abn}</p>
<p class="sub">{period_str}</p>

<h2>GST on Sales</h2>
<table>
    <tr><th>Label</th><th>Description</th><th class="r">Amount</th></tr>
    <tr><td class="bold">G1</td><td>Total sales (including any GST)</td><td class="r">{fmt(bas['G1'])}</td></tr>
    <tr><td>G2</td><td>Export sales</td><td class="r">{fmt(bas['G2'])}</td></tr>
    <tr><td>G3</td><td>Other GST-free sales</td><td class="r">{fmt(bas['G3'])}</td></tr>
    <tr><td>G4</td><td>Input taxed sales</td><td class="r">{fmt(bas['G4'])}</td></tr>
    <tr class="summary"><td class="bold">G5</td><td>G2 + G3 + G4</td><td class="r bold">{fmt(bas['G5'])}</td></tr>
    <tr><td class="bold">G6</td><td>Total sales subject to GST (G1 &minus; G5)</td><td class="r bold">{fmt(bas['G6'])}</td></tr>
    <tr><td>G7</td><td>Adjustments</td><td class="r">{fmt(bas['G7'])}</td></tr>
    <tr class="summary"><td class="bold">G8</td><td>Total sales subject to GST after adj. (G6 + G7)</td><td class="r bold">{fmt(bas['G8'])}</td></tr>
    <tr class="highlight"><td class="bold">G9</td><td>GST on sales (G8 &divide; 11)</td><td class="r bold">{fmt(bas['G9'])}</td></tr>
</table>

<h2>GST on Purchases</h2>
<table>
    <tr><th>Label</th><th>Description</th><th class="r">Amount</th></tr>
    <tr><td class="bold">G10</td><td>Capital purchases (including any GST)</td><td class="r">{fmt(bas['G10'])}</td></tr>
    <tr><td class="bold">G11</td><td>Non-capital purchases (including any GST)</td><td class="r">{fmt(bas['G11'])}</td></tr>
    <tr class="summary"><td class="bold">G12</td><td>G10 + G11</td><td class="r bold">{fmt(bas['G12'])}</td></tr>
    <tr><td>G13</td><td>Purchases for making input taxed sales</td><td class="r">{fmt(bas['G13'])}</td></tr>
    <tr><td>G14</td><td>Purchases without GST in the price</td><td class="r">{fmt(bas['G14'])}</td></tr>
    <tr><td>G15</td><td>Estimated purchases for private use</td><td class="r">{fmt(bas['G15'])}</td></tr>
    <tr class="summary"><td class="bold">G16</td><td>G13 + G14 + G15</td><td class="r bold">{fmt(bas['G16'])}</td></tr>
    <tr><td class="bold">G17</td><td>Total purchases subject to GST (G12 &minus; G16)</td><td class="r bold">{fmt(bas['G17'])}</td></tr>
    <tr><td>G18</td><td>Adjustments</td><td class="r">{fmt(bas['G18'])}</td></tr>
    <tr class="summary"><td class="bold">G19</td><td>Total purchases subject to GST after adj. (G17 + G18)</td><td class="r bold">{fmt(bas['G19'])}</td></tr>
    <tr class="highlight"><td class="bold">G20</td><td>GST on purchases (G19 &divide; 11)</td><td class="r bold">{fmt(bas['G20'])}</td></tr>
</table>

<h2>Activity Statement Summary</h2>
<table style="max-width: 400px;">
    <tr><td class="bold">1A</td><td>GST on sales</td><td class="r bold">{fmt(bas['1A'])}</td></tr>
    <tr><td class="bold">1B</td><td>GST on purchases (credit)</td><td class="r bold">{fmt(bas['1B'])}</td></tr>
    <tr class="total-row"><td colspan="2">{net_label}</td><td class="r">{fmt(abs(net_gst))}</td></tr>
</table>

<h2>Detail Breakdown</h2>
<table>
    <tr><th>Code</th><th>Account Name</th><th>Tax Code</th><th class="r">Amount</th><th>BAS Label</th></tr>
    {detail_html}
</table>

{lodgement_footer}

</body></html>"""

    pdf_bytes = weasyprint.HTML(string=html).write_pdf()

    safe_name = entity.entity_name.replace(' ', '_')
    safe_period = period_label.replace(' ', '_').replace('(', '').replace(')', '')
    if period_start and period_end:
        filename = f"BAS_{safe_name}_{safe_period}_{period_start.strftime('%Y%m%d')}_{period_end.strftime('%Y%m%d')}.pdf"
    else:
        filename = f"BAS_{safe_name}_FullYear_{fy.start_date.strftime('%Y%m%d')}_{fy.end_date.strftime('%Y%m%d')}.pdf"

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _download_all_periods(fy, entity, period_type, fmt):
    """
    Generate a multi-period download.
    Excel: one sheet per period + Full Year summary sheet.
    PDF: multi-page with page breaks between periods.
    """
    periods = ensure_bas_periods(fy, period_type)

    if fmt == "pdf":
        return _all_periods_pdf(fy, entity, periods)

    return _all_periods_excel(fy, entity, periods)


def _all_periods_excel(fy, entity, periods):
    """Multi-sheet Excel with one sheet per period + Full Year."""
    wb = openpyxl.Workbook()

    # Full Year sheet first
    ws = wb.active
    full_year_result = calculate_gst_for_period(fy, None, None)
    full_year_detail = _flatten_detail_rows(full_year_result)
    _write_bas_sheet(ws, fy, entity, "Full Year", full_year_result["bas_data"], full_year_detail)

    # Per-period sheets
    for bp in periods:
        result = calculate_gst_for_period(fy, bp.period_start, bp.period_end)
        detail = _flatten_detail_rows(result)
        ws_p = wb.create_sheet(bp.short_label)
        _write_bas_sheet(ws_p, fy, entity, bp.label, result["bas_data"], detail, bp.period_start, bp.period_end)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_name = entity.entity_name.replace(" ", "_")
    filename = f"BAS_{safe_name}_AllPeriods_{fy.start_date.strftime('%Y%m%d')}_{fy.end_date.strftime('%Y%m%d')}.xlsx"

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _all_periods_pdf(fy, entity, periods):
    """Multi-page PDF with page breaks between periods."""
    import weasyprint
    from django.utils.html import escape as html_escape

    abn = html_escape(entity.abn or 'N/A')

    def fmt(v):
        return f"${v:,.0f}"

    pages = []

    # Full Year page
    full_year_result = calculate_gst_for_period(fy, None, None)
    pages.append(_build_pdf_page(
        entity, abn, "Full Year",
        f"{fy.start_date.strftime('%d/%m/%Y')} to {fy.end_date.strftime('%d/%m/%Y')}",
        full_year_result["bas_data"],
        _flatten_detail_rows(full_year_result),
        fmt,
    ))

    # Per-period pages
    for bp in periods:
        result = calculate_gst_for_period(fy, bp.period_start, bp.period_end)
        period_str = f"{bp.period_start.strftime('%d/%m/%Y')} to {bp.period_end.strftime('%d/%m/%Y')}"
        pages.append(_build_pdf_page(
            entity, abn, bp.label, period_str,
            result["bas_data"],
            _flatten_detail_rows(result),
            fmt,
        ))

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
    @page {{ size: A4; margin: 15mm; }}
    body {{ font-family: 'Times New Roman', serif; font-size: 10pt; color: #333; }}
    h1 {{ font-size: 14pt; text-align: center; margin-bottom: 2mm; }}
    h2 {{ font-size: 11pt; margin-top: 6mm; margin-bottom: 3mm; border-bottom: 1px solid #333; padding-bottom: 2mm; }}
    .sub {{ text-align: center; font-size: 10pt; color: #666; margin-bottom: 4mm; }}
    table {{ width: 100%; border-collapse: collapse; margin-bottom: 4mm; }}
    th, td {{ padding: 3px 6px; font-size: 9pt; border-bottom: 1px solid #ddd; }}
    th {{ background: #f5f5f5; text-align: left; font-weight: bold; }}
    .r {{ text-align: right; }}
    .bold {{ font-weight: bold; }}
    .highlight {{ background: #e8f4fd; }}
    .summary {{ background: #f8f9fa; }}
    .total-row {{ border-top: 2px solid #333; font-weight: bold; }}
    .page-break {{ page-break-before: always; }}
</style></head><body>
{''.join(pages)}
</body></html>"""

    pdf_bytes = weasyprint.HTML(string=html).write_pdf()

    safe_name = entity.entity_name.replace(' ', '_')
    filename = f"BAS_{safe_name}_AllPeriods_{fy.start_date.strftime('%Y%m%d')}_{fy.end_date.strftime('%Y%m%d')}.pdf"

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ── Helper Functions ─────────────────────────────────────────────────────────

def _flatten_detail_rows(gst_result):
    """Flatten sales/purchase/capital lines into a single detail list."""
    rows = []
    for lines in [gst_result["sales_lines"], gst_result["purchase_lines"], gst_result["capital_lines"]]:
        for row in lines:
            rows.append({
                "code": row["code"],
                "name": row["name"],
                "tax_code": row.get("tax_code", "N-T"),
                "amount": float(row["amount"]),
                "bas_label": row.get("bas_label", ""),
            })
    return rows


def _write_bas_sheet(ws, fy, entity, period_label, bas, detail_rows, period_start=None, period_end=None):
    """Write BAS summary data to an Excel worksheet."""
    ws.title = f"BAS - {period_label}"[:31]  # Excel sheet name max 31 chars

    bold = Font(bold=True)

    # Header
    ws.append([f"GST Activity Statement — {entity.entity_name}"])
    if period_start and period_end:
        ws.append([f"Period: {period_label} ({period_start.strftime('%d/%m/%Y')} to {period_end.strftime('%d/%m/%Y')})"])
    else:
        ws.append([f"Period: Full Year ({fy.start_date.strftime('%d/%m/%Y')} to {fy.end_date.strftime('%d/%m/%Y')})"])
    ws.append([f"ABN: {entity.abn or 'N/A'}"])
    ws.append([])

    # GST on Sales
    ws.append(["GST ON SALES"])
    ws.append(["Label", "Description", "Amount"])
    ws.append(["G1", "Total sales (including any GST)", float(bas["G1"])])
    ws.append(["G2", "Export sales", float(bas["G2"])])
    ws.append(["G3", "Other GST-free sales", float(bas["G3"])])
    ws.append(["G4", "Input taxed sales", float(bas["G4"])])
    ws.append(["G5", "G2 + G3 + G4", float(bas["G5"])])
    ws.append(["G6", "Total sales subject to GST (G1 - G5)", float(bas["G6"])])
    ws.append(["G7", "Adjustments", float(bas["G7"])])
    ws.append(["G8", "Total sales subject to GST after adjustments (G6 + G7)", float(bas["G8"])])
    ws.append(["G9", "GST on sales (G8 / 11)", float(bas["G9"])])
    ws.append([])

    # GST on Purchases
    ws.append(["GST ON PURCHASES"])
    ws.append(["Label", "Description", "Amount"])
    ws.append(["G10", "Capital purchases (including any GST)", float(bas["G10"])])
    ws.append(["G11", "Non-capital purchases (including any GST)", float(bas["G11"])])
    ws.append(["G12", "G10 + G11", float(bas["G12"])])
    ws.append(["G13", "Purchases for making input taxed sales", float(bas["G13"])])
    ws.append(["G14", "Purchases without GST in the price", float(bas["G14"])])
    ws.append(["G15", "Estimated purchases for private use", float(bas["G15"])])
    ws.append(["G16", "G13 + G14 + G15", float(bas["G16"])])
    ws.append(["G17", "Total purchases subject to GST (G12 - G16)", float(bas["G17"])])
    ws.append(["G18", "Adjustments", float(bas["G18"])])
    ws.append(["G19", "Total purchases subject to GST after adjustments (G17 + G18)", float(bas["G19"])])
    ws.append(["G20", "GST on purchases (G19 / 11)", float(bas["G20"])])
    ws.append([])

    # Summary
    ws.append(["BAS SUMMARY"])
    ws.append(["Label", "Description", "Amount"])
    ws.append(["1A", "GST on sales", float(bas["1A"])])
    ws.append(["1B", "GST on purchases", float(bas["1B"])])
    ws.append(["", "Net GST payable / (refundable)", float(bas["gst_payable"])])

    # Format columns
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 18

    # Bold headers
    for row_idx in [1, 2, 3, 5, 6, 17, 18, 31, 32]:
        if row_idx <= ws.max_row:
            for cell in ws[row_idx]:
                cell.font = bold

    # Number format
    for row in ws.iter_rows(min_row=7, max_col=3, max_row=ws.max_row):
        cell = row[2]
        if isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0'


def _write_detail_sheet(ws, detail_rows):
    """Write detail breakdown to an Excel worksheet."""
    bold = Font(bold=True)
    ws.append(["Account Code", "Account Name", "Tax Code", "Amount", "BAS Label"])
    for row in detail_rows:
        ws.append([row["code"], row["name"], row["tax_code"], row["amount"], row["bas_label"]])

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14

    for cell in ws[1]:
        cell.font = bold
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4, max_row=ws.max_row):
        row[0].number_format = '#,##0'


def _build_pdf_page(entity, abn, period_label, period_str, bas, detail_rows, fmt, is_first=False):
    """Build HTML for a single page of the multi-period PDF."""
    from django.utils.html import escape as html_escape

    net_gst = bas["gst_payable"]
    net_label = "Net GST Payable to ATO" if net_gst > 0 else "Net GST Refundable from ATO"

    detail_html = ""
    for row in detail_rows:
        detail_html += f"""<tr>
            <td>{html_escape(str(row['code']))}</td><td>{html_escape(str(row['name']))}</td>
            <td>{html_escape(str(row['tax_code']))}</td><td class="r">${row['amount']:,.0f}</td>
            <td>{html_escape(str(row['bas_label']))}</td>
        </tr>"""

    page_break = '' if is_first else ' class="page-break"'

    return f"""
<div{page_break}>
<h1>GST Activity Statement — {period_label}</h1>
<p class="sub">{html_escape(entity.entity_name)} &mdash; ABN: {abn}</p>
<p class="sub">{period_str}</p>

<h2>GST on Sales</h2>
<table>
    <tr><th>Label</th><th>Description</th><th class="r">Amount</th></tr>
    <tr><td class="bold">G1</td><td>Total sales (including any GST)</td><td class="r">{fmt(bas['G1'])}</td></tr>
    <tr><td>G2</td><td>Export sales</td><td class="r">{fmt(bas['G2'])}</td></tr>
    <tr><td>G3</td><td>Other GST-free sales</td><td class="r">{fmt(bas['G3'])}</td></tr>
    <tr><td>G4</td><td>Input taxed sales</td><td class="r">{fmt(bas['G4'])}</td></tr>
    <tr class="summary"><td class="bold">G5</td><td>G2 + G3 + G4</td><td class="r bold">{fmt(bas['G5'])}</td></tr>
    <tr><td class="bold">G6</td><td>Total sales subject to GST (G1 &minus; G5)</td><td class="r bold">{fmt(bas['G6'])}</td></tr>
    <tr><td>G7</td><td>Adjustments</td><td class="r">{fmt(bas['G7'])}</td></tr>
    <tr class="summary"><td class="bold">G8</td><td>Total sales subject to GST after adj. (G6 + G7)</td><td class="r bold">{fmt(bas['G8'])}</td></tr>
    <tr class="highlight"><td class="bold">G9</td><td>GST on sales (G8 &divide; 11)</td><td class="r bold">{fmt(bas['G9'])}</td></tr>
</table>

<h2>GST on Purchases</h2>
<table>
    <tr><th>Label</th><th>Description</th><th class="r">Amount</th></tr>
    <tr><td class="bold">G10</td><td>Capital purchases (including any GST)</td><td class="r">{fmt(bas['G10'])}</td></tr>
    <tr><td class="bold">G11</td><td>Non-capital purchases (including any GST)</td><td class="r">{fmt(bas['G11'])}</td></tr>
    <tr class="summary"><td class="bold">G12</td><td>G10 + G11</td><td class="r bold">{fmt(bas['G12'])}</td></tr>
    <tr><td>G13</td><td>Purchases for making input taxed sales</td><td class="r">{fmt(bas['G13'])}</td></tr>
    <tr><td>G14</td><td>Purchases without GST in the price</td><td class="r">{fmt(bas['G14'])}</td></tr>
    <tr><td>G15</td><td>Estimated purchases for private use</td><td class="r">{fmt(bas['G15'])}</td></tr>
    <tr class="summary"><td class="bold">G16</td><td>G13 + G14 + G15</td><td class="r bold">{fmt(bas['G16'])}</td></tr>
    <tr><td class="bold">G17</td><td>Total purchases subject to GST (G12 &minus; G16)</td><td class="r bold">{fmt(bas['G17'])}</td></tr>
    <tr><td>G18</td><td>Adjustments</td><td class="r">{fmt(bas['G18'])}</td></tr>
    <tr class="summary"><td class="bold">G19</td><td>Total purchases subject to GST after adj. (G17 + G18)</td><td class="r bold">{fmt(bas['G19'])}</td></tr>
    <tr class="highlight"><td class="bold">G20</td><td>GST on purchases (G19 &divide; 11)</td><td class="r bold">{fmt(bas['G20'])}</td></tr>
</table>

<h2>Activity Statement Summary</h2>
<table style="max-width: 400px;">
    <tr><td class="bold">1A</td><td>GST on sales</td><td class="r bold">{fmt(bas['1A'])}</td></tr>
    <tr><td class="bold">1B</td><td>GST on purchases (credit)</td><td class="r bold">{fmt(bas['1B'])}</td></tr>
    <tr class="total-row"><td colspan="2">{net_label}</td><td class="r">{fmt(abs(net_gst))}</td></tr>
</table>

{f'<h2>Detail Breakdown</h2><table><tr><th>Code</th><th>Account Name</th><th>Tax Code</th><th class="r">Amount</th><th>BAS Label</th></tr>{detail_html}</table>' if detail_rows else ''}
</div>"""
